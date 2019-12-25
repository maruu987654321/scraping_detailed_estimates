import sqlalchemy
import mysql.connector
import pymysql
import yaml
import pandas as pd 
from openpyxl.styles import Alignment
from openpyxl import load_workbook

def from_df_to_excel(name):
    database_connection = sqlalchemy.create_engine('mysql+mysqlconnector://{0}:{1}@{2}/{3}'.
                                                    format(database_username, database_password, 
                                                        database_ip, database_name), pool_recycle=1, pool_timeout=57600).connect()
    df = pd.read_sql('SELECT * FROM {}'.format(name), con=database_connection)
    return df

def load_config(config_file):
    with open(config_file, 'r') as stream:
        try:
            return yaml.safe_load(stream)
        except yaml.YAMLError as exc:
            print(exc)

if __name__ == '__main__':
    config = load_config('config.yaml')
    database_username = config['database_username']   #type your username
    database_password =  config['password']   #type your password
    database_ip       =  config['database_ip']
    database_name     = config['name_database']  #type name your database
    name_file_with_stock = config['name_file_with_stocks'] #name of file where lie all stocks
    mydb = mysql.connector.connect(
                host=database_ip,
                user=database_username,
                passwd=database_password,
                )

    mycursor = mydb.cursor()
    mycursor.execute("USE {}".format(database_name))
    mycursor.execute("SHOW TABLES")
    tables = mycursor.fetchall()
    for i in tables:
        str_n = ''.join(i)
        df = from_df_to_excel(str_n)
        try:
            del df['index']
            writer = pd.ExcelWriter('{}.xlsx'.format(str_n))
            df.to_excel(writer, '{}'.format(str_n),  index=False)
            writer.save()
            wb = load_workbook("{}.xlsx".format(str_n))
            ws = wb['{}'.format(str_n)]
            al = Alignment(horizontal='right')
            ws.column_dimensions['A'].width = 30 #width for A coolumn
            ws.column_dimensions['B'].width = 25 #width for B column
            ws.column_dimensions['C'].width = 25 #width for C column
            ws.column_dimensions['D'].width = 25 #width for D column
            ws.column_dimensions['E'].width = 25  #width for E column
            for i in range(1, 51):
                first_row = 'B{}'.format(i)
                second_row = 'Y{}'.format(i)
                for row in ws[first_row:second_row]: 
                    for cell in row: 
                        cell.alignment = al 

            wb.save(filename='{}.xlsx'.format(str_n))

        except KeyError:
            df.to_excel('{}.xlsx'.format(str_n), engine='xlsxwriter', index=False)